from PyPDF2 import PdfReader, PdfWriter
from reportlab.lib.pagesizes import A4
from reportlab.pdfgen import canvas
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.platypus import Paragraph
import pandas as pd
from openpyxl import load_workbook, Workbook
from openpyxl.utils import get_column_letter
import io, os, sys
from io import BytesIO
from pathlib import Path
from PIL import Image  
import pdf_data, configparser, canvas_update
from datetime import datetime
import tkinter.messagebox as messagebox
from num2words import num2words
import locale, re

# Load configurations
config_path = os.path.join(os.path.dirname(__file__), "config", "receipt.ini")
config = configparser.ConfigParser()
config.read(config_path)
img_path = os.path.join(os.path.dirname(__file__), "Images", "Logo.jpg")
font_name = config.get('FontSettings', 'font_name')
copy_font_name = config.get('FontSettings', 'copy_font_name')
font_name_bold = config.get('FontSettings', 'font_name_bold')
#font_name_bold = config.get('FontSettings', 'font_name_bold')
font_size = config.getint('FontSettings', 'font_size')
font_size_bold = config.getint('FontSettings', 'font_size_bold')
xcl_sheet = config.get('Filenames', 'xcl_sheet')
width, height = A4
height = height - 56
x_offset = 90
headings = {key: config.get('Heading', key) for key in ['heading', 'sub_heading', 'inv_heading']}
y_position = config.getint('FontSettings', 'y_position')
x_v_date = config.getint('voucher_dimensions', 'x_v_date')
y_top_v_date = config.get('voucher_dimensions', 'y_top_v_date')
y_middle_v_date = config.get('voucher_dimensions', 'y_middle_v_date')
y_bottom_v_date = config.get('voucher_dimensions', 'y_bottom_v_date')

x_amount = config.getint('voucher_dimensions', 'x_amount')
y_top_amount = config.get('voucher_dimensions', 'y_top_amount')
y_middle_amount = config.get('voucher_dimensions', 'y_middle_amount')
y_bottom_amount = config.get('voucher_dimensions', 'y_bottom_amount')

x_amount_in_words = config.getint('voucher_dimensions', 'x_amount_in_words')
y_top_amount_in_words = config.get('voucher_dimensions', 'y_top_amount_in_words')
y_middle_amount_in_words = config.get('voucher_dimensions', 'y_middle_amount_in_words')
y_bottom_amount_in_words = config.get('voucher_dimensions', 'y_bottom_amount_in_words')
line_width = config.getint('voucher_dimensions', 'line_width')

x_payto_pan = config.getint('voucher_dimensions', 'x_payto_pan')
y_top_payto = config.get('voucher_dimensions', 'y_top_payto')
y_middle_payto = config.get('voucher_dimensions', 'y_middle_payto')
y_bottom_payto = config.get('voucher_dimensions', 'y_bottom_payto')

y_top_pan= config.get('voucher_dimensions', 'y_top_pan')
y_middle_pan = config.get('voucher_dimensions', 'y_middle_pan')
y_bottom_pan = config.get('voucher_dimensions', 'y_bottom_pan')

x_qrcode = config.getint('voucher_dimensions', 'x_qrcode')
y_start_qrcode = config.get('voucher_dimensions', 'y_start_qrcode')
y_offset = config.getint('voucher_dimensions', 'y_offset')
x_inv_qrcode = config.getint('invoice_dimensions', 'x_inv_qrcode')
y_inv_start_qrcode = config.getint('invoice_dimensions', 'y_inv_start_qrcode')
inv_qrcode_size = config.getint('invoice_dimensions', 'inv_qrcode_size')
x_logo = config.getint('invoice_dimensions', 'x_logo')

x_inv_barcode = config.getint('invoice_dimensions', 'x_inv_barcode')
y_inv_start_barcode = config.getint('invoice_dimensions', 'y_inv_start_barcode')
inv_barcode_width = config.getint('invoice_dimensions', 'inv_barcode_width')
inv_barcode_height = config.getint('invoice_dimensions', 'inv_barcode_height')

x_position = config.getint('invoice_dimensions', 'x_position')

x_authoriser = config.getint('voucher_dimensions', 'x_authoriser')
y_top_authoriser = config.get('voucher_dimensions', 'y_top_authoriser')
y_middle_authoriser = config.get('voucher_dimensions', 'y_middle_authoriser')
y_bottom_authoriser = config.get('voucher_dimensions', 'y_bottom_authoriser')

y_top_print_date = config.get('voucher_dimensions', 'y_top_print_date')
y_middle_print_date = config.get('voucher_dimensions', 'y_middle_print_date')
y_bottom_print_date = config.get('voucher_dimensions', 'y_bottom_print_date')

x_inv_date = config.getint('invoice_dimensions', 'x_inv_date')
y_inv_date = config.get('invoice_dimensions', 'y_inv_date')

# Converting date to the desired format
def get_date_format(c, value):
    date_obj = datetime.strptime(value, "%d/%m/%Y") 
    formatted_date = date_obj.strftime("%d%b%Y").upper()
    c.setFont(font_name, font_size) 
    return formatted_date

# Function to draw voucher date inside the boxes
def draw_voucher_date(c, i, value):
    formatted_date = get_date_format(c, value)
    c.setFont(font_name, font_size) 

    if i == 0:
        c.drawString(x_v_date, eval(y_top_v_date), "    ".join(formatted_date) )
    elif i == 1:
        c.drawString(x_v_date, eval(y_middle_v_date), "    ".join(formatted_date) )
    elif i == 2:
        c.drawString(x_v_date, eval(y_bottom_v_date), "    ".join(formatted_date) )
    # elif i == 3:
    #     c.drawString(x_inv_date, eval(y_inv_date), formatted_date )
    #c.setFont(font_name, font_size) 


def print_amount_ac(c, locale_value, x1, x2, y1, y2, combined_amount_ac, line_width):
    if locale_value != "":
            c.drawString(x1, y1, f"{locale_value} /-")
    wrap_text_voucher(c, f"{combined_amount_ac}", x2, y2, line_width, font_name, font_size, 1)

# Function to wrap text
def wrap_text_voucher(c, text, x, y, width, font_name, font_size, receipt):
    styles = getSampleStyleSheet()
    style = styles['Normal']
    style.fontName = font_name
    style.fontSize = font_size
    style.leading = font_size + 9                           #vertical spacing between lines
    
    # To adjust the max 200 chars in Address 
    if receipt:
        style.leftIndent = 0
        style.firstLineIndent = 80
    else:
        style.leftIndent = 0
        style.firstLineIndent = 40
    #    #style.fontSize = font_size-4

    #if "&nbsp;&nbsp;&nbsp;&nbsp;" not in text:              #vertical spacing for 'Amount in words' is different
    #    style.leading = font_size + 8

    paragraph = Paragraph(text, style)                      # Create a Paragraph object which wraps the text
    _, text_height = paragraph.wrap(width, 600)             # Get actual text height
    paragraph.drawOn(c, x-80, y - text_height)     

# Custom function to convert number to INR words (Indian Rupees in English)
def convert_to_words(amount):
    try:
        amount = int(amount)
    except ValueError:
        #messagebox.showerror("Amount Error", "Please enter a valid amount.")
        return
    
    whole_number = int(amount)                              # Split the amount into whole and fractional parts
    fraction_part = round((amount - whole_number) * 100)    # Getting the fraction part as Paise
    whole_in_words = num2words(whole_number, lang='en_IN')  # Convert whole number into words in English
    capitalized_number = whole_in_words.upper()             # Capitalize the first letter of each word
    if fraction_part > 0:                                   # Convert fraction part (Paise) into words if it's non-zero
        fraction_in_words = num2words(fraction_part, lang='en_IN')
        capitalized_fractions = fraction_in_words.upper()
        result = f"{capitalized_number} RUPEES {capitalized_fractions} PAISE"
    else:
        result = f"{capitalized_number} RUPEES"
    return result

def underline_text(canvas, x_pos, y_pos, text):
    canvas.setLineWidth(0.5)  # Set line thickness
    if text in headings or text in headings['inv_heading']:
        text_width = canvas.stringWidth(text, font_name, font_size)
    else:
        text_width = canvas.stringWidth(text, font_name, font_size)
    canvas.line(x_pos, y_pos - 2, x_pos + (text_width), y_pos - 2)

def center_text(canvas, text, y_pos, fonts):
    page_width, page_height = A4
    if not isinstance(text, str):
        text = str(text)
    text_width = canvas.stringWidth(text, font_name_bold, font_size_bold)
    x_pos = (page_width - text_width) / 2  # Center the text horizontally
    canvas.setFont(font_name_bold, font_size_bold)
    canvas.drawString(x_pos, y_pos, text)
    underline_text(canvas, x_pos, y_pos, text)

def draw_print_date(c, y, value):
    c.setFillColorRGB(0.5, 0.5, 0.5) 
    c.setFont(copy_font_name, 8)  # Set font and size
    canvas_update.draw_copy_type(c, y, value)
    c.setFont(font_name, font_size)
    return

# Function to draw amount in numbers and amount in words
def draw_voucher(c, i, value, pay_to, pan, addr, pur_code, pur_head, pur_cat, exp_type, mode, inst_date, inst_no_val, IFSC, inst_AC_no, bank_name, bank_branch, dest_excel_path, pdf_path, user, print_date, CIN):
    c.setFont(font_name, font_size)
    locale.setlocale(locale.LC_ALL, 'en_IN')
    if value:
        locale_value = locale.format_string("%d", int(value), grouping=True)
    else:
        locale_value = ""
        
    in_words = convert_to_words(value)
    if in_words is None:
        in_words = ""
    else:
        in_words = in_words.title() + " Only ; "

    combined_payto = pay_to #+ ". PAN: " + pan + " Address: " + addr
    if pan:
        combined_payto += f". PAN: {pan}"
    if addr:
        combined_payto += f" Address: {addr}"
    
    if CIN and CIN != ' ':
        CIN = "TDS Stat(CIN):" + CIN
    else:
        CIN = ''

    if pur_code:
        combined_purchase = "PC: " + pur_code + " PH: " + pur_head + " PCat: " + pur_cat + " EP: " + exp_type + CIN
    else:
        combined_purchase = '' + CIN

    if mode == 'Cheque':
        combined_amount_ac = in_words + " CH Date:" + inst_date + " CH#:" + inst_no_val + " IFSC:" + IFSC + " A/C#:" + inst_no_val
    elif mode == 'EFT':
        combined_amount_ac = in_words + " EFT Date:" + inst_date + " UTRN:" + inst_no_val + " Bank Name:" + bank_name + "Bank Branch:" + bank_branch
    else:
        combined_amount_ac = in_words

    if user is None:
        user = ''
    print_date = f"{user} {print_date}"

    if i == 0:
        print_amount_ac(c, locale_value, x_amount, x_amount_in_words, eval(y_top_amount), eval(y_top_amount_in_words), combined_amount_ac, line_width)
        wrap_text_voucher(c, f"{combined_payto}", x_payto_pan, eval(y_top_payto), line_width, font_name, font_size, 0)
        wrap_text_voucher(c, f"{combined_purchase}", x_payto_pan, eval(y_top_pan), line_width, font_name, font_size, 0)
        draw_print_date(c, eval(y_top_print_date), print_date)
        #c.drawString(x_authoriser, eval(y_top_authoriser), f"{user}")
        c.setFillColorRGB(0, 0, 0) 
    elif i == 1:
        print_amount_ac(c, locale_value, x_amount, x_amount_in_words, eval(y_middle_amount), eval(y_middle_amount_in_words), combined_amount_ac, line_width)
        wrap_text_voucher(c, f"{combined_payto}", x_payto_pan, eval(y_middle_payto), line_width, font_name, font_size, 0)
        wrap_text_voucher(c, f"{combined_purchase}", x_payto_pan, eval(y_middle_pan), line_width, font_name, font_size, 0)
        draw_print_date(c, eval(y_middle_print_date), print_date)
        #c.drawString(x_authoriser, eval(y_middle_authoriser), f"{user}")
        c.setFillColorRGB(0, 0, 0) 
    elif i == 2:
        print_amount_ac(c, locale_value, x_amount, x_amount_in_words, eval(y_bottom_amount), eval(y_bottom_amount_in_words), combined_amount_ac, line_width)
        wrap_text_voucher(c, f"{combined_payto}", x_payto_pan, eval(y_bottom_payto), line_width, font_name, font_size, 0)
        wrap_text_voucher(c, f"{combined_purchase}", x_payto_pan, eval(y_bottom_pan), line_width, font_name, font_size, 0)
        draw_print_date(c, eval(y_bottom_print_date), print_date)
        #c.drawString(x_authoriser, eval(y_bottom_authoriser), f"{user}")
        c.setFillColorRGB(0, 0, 0) 
    
    draw_code(c, dest_excel_path)
    return True

def draw_code(c, dest_excel_path):
    print(f"DEST path {dest_excel_path}")
    x_qrcode_DN_DPS = config.getint('voucher_dimensions', 'x_qrcode_DN_DPS')
    x_qrcode_SGPT = config.getint('voucher_dimensions', 'x_qrcode_SGPT')
    qrcode_width = config.getint('voucher_dimensions', 'qrcode_width')
    qrcode_height = config.getint('voucher_dimensions', 'qrcode_height')

    x_barcode = config.getint('voucher_dimensions', 'x_barcode')
    y_start_barcode = config.get('voucher_dimensions', 'y_start_barcode')
    barcode_width = config.getint('voucher_dimensions', 'barcode_width')
    barcode_height = config.getint('voucher_dimensions', 'barcode_height')

    wb = load_workbook(dest_excel_path)
    ws = wb[xcl_sheet]    #wb.active

    # Get header row to find correct columns
    header_row = next(ws.iter_rows(min_row=2, max_row=2, values_only=True))
    #col_map = {name.strip().lower(): idx for idx, name in enumerate(header_row) if name is not None}

    # Ensure columns exist
    barcode_col_idx = header_row.index("Bar Code")
    qrcode_col_idx = header_row.index("QR Code")
    if barcode_col_idx is None or qrcode_col_idx is None:
        raise ValueError("Required columns 'Bar Code' and/or 'QRCode' not found.")

    barcode_col_letter = get_column_letter(barcode_col_idx + 1)
    qrcode_col_letter = get_column_letter(qrcode_col_idx + 1)

    # Separate images by column and row
    barcode_images = {}
    qrcode_images = {}

    #print(len(ws._images))

    for img in ws._images:
        if hasattr(img.anchor, "_from"):
            col = img.anchor._from.col
            row = img.anchor._from.row
            col_letter = get_column_letter(col + 2)

            if col_letter == barcode_col_letter:
                barcode_images[row] = img
            elif col_letter == qrcode_col_letter:
                qrcode_images[row] = img

    # Find all rows where at least a barcode or QR code exists
    all_rows = sorted(set(barcode_images.keys()) | set(qrcode_images.keys()))

    if 'invoice' in dest_excel_path:
        last_rows = all_rows[-1:] 
        x_barcode = x_inv_barcode
        y_barcode = y_inv_start_barcode
        qrcode_width = inv_qrcode_size
        qrcode_height = inv_qrcode_size
        barcode_width = inv_barcode_width
        barcode_height = inv_barcode_height
        x_qrcode_DN_DPS = x_qrcode_SGPT = x_qrcode = x_inv_qrcode
    else:
        last_rows = all_rows[-3:]                                # Get last 3 rows with image(s)
                                              
    if 'SGPM_DN' in dest_excel_path or 'SPK_DPS' in dest_excel_path:
        x_qrcode_updated = x_qrcode_DN_DPS
    elif 'SGPT' in dest_excel_path:
        x_qrcode_updated = x_qrcode_SGPT
    else:
        x_qrcode_updated = x_qrcode
    
    # Draw images row-wise
    for i, row in enumerate(last_rows):
        if 'invoice' in dest_excel_path:
            y_barcode = y_inv_start_barcode
            y_qrcode = y_inv_start_qrcode
        else:
            y_barcode = eval(y_start_barcode) - i * y_offset
            y_qrcode = eval(y_start_qrcode) - i * y_offset

        # Bar Code
        if row in barcode_images:
            img_data = barcode_images[row]._data()
            pil_img = Image.open(BytesIO(img_data))
            pil_img.save(f"temp_barcode_{row}.png")
            c.drawImage(f"temp_barcode_{row}.png", x_barcode, y_barcode, width=barcode_width, height=barcode_height)
            os.remove(f"temp_barcode_{row}.png")
        # QRCode
        if row in qrcode_images:
            img_data = qrcode_images[row]._data()
            pil_img = Image.open(BytesIO(img_data))
            pil_img.save(f"temp_qrcode_{row}.png")
            c.drawImage(f"temp_qrcode_{row}.png", x_qrcode_updated, y_qrcode, width=qrcode_width, height=qrcode_height)
            os.remove(f"temp_qrcode_{row}.png")

def print_label_value(c, x_logo, x_position, y_position, label, value):
    c.drawString(x_logo, y_position, label)
    underline_text(c, x_logo, y_position, label)
    c.drawString(x_position, y_position, str(value))
    y_position -= 15
    return y_position

#def draw_invoice(c, amount, pay_to, addr, pur_code, pur_head, pur_cat, exp_type, mode, cheque_date, cheque_no_val, IFSC, cheque_AC_no, eft_date, utrn_val, bank_name, bank_branch, CIN, dest_excel_path, selected_idx):
def draw_invoice(c, voucher_date, amount, pay_to, addr, pur_code, pur_head, pur_cat, exp_type, mode, inst_date, inst_no_val, IFSC, inst_AC_no, bank_name, bank_branch, CIN, dest_excel_path, selected_idx, inv_no, trans_type):
    print("PRINT INVOICE HERE")
    y_position = 550
    text = headings['heading'].split(',')
    fields = [
    ("Invoice # :", inv_no),
    ("Invoice Date:", voucher_date),
    ("Pay to:", pay_to),
    ("Amount:", amount),
    ("Address:", addr),
    ("Purchase Code:", pur_code),
    ("Purchase Head:", pur_head),
    ("Purchase Category:", pur_cat),
    ("Expense Type:", exp_type),
    ("CIN:", CIN),
    ("Payment Mode:", mode),
    ("Transaction Type:", trans_type),
]
    fields_mode = [ 
    ("Payment Mode:", mode),
    ("Cheque Date:", inst_date),
    ("Cheque No:", inst_no_val),
    ("Account No:", inst_AC_no),
    ("EFT Date:", inst_date),
    ("UTR No:", inst_no_val),
    ("IFSC:", IFSC),
    ("Bank Name:", bank_name),
    ("Bank Branch:", bank_branch)
    ]
    center_text(c, text[selected_idx], y_position, font_name)
    c.setLineWidth(3)   # Set line thickness (3 is a thick line, you can adjust it)
    c.line(50, y_position - 5, 550, y_position - 5)
    y_position -= 20
    center_text(c, str(headings['sub_heading']), y_position, font_name)
    y_position -= 20
    #center_text(c, str(headings['inv_heading']), y_position, font_name)
    #y_position -= 20

    draw_code(c, dest_excel_path)

    c.drawImage(img_path, x_logo, y_inv_start_qrcode, width=100, height=100, mask=None)

    c.setFont(font_name, font_size) 
    for label, value in fields:
        #if (value) and value != ' ':
        if label == 'Amount:':
            locale.setlocale(locale.LC_ALL, 'en_IN.UTF-8')
            value = int(value)
            value = locale.format_string("%.0f", value, grouping=True)
        elif label == 'Address:':
            value = value.replace('\n', ' ')
        y_position = print_label_value(c, x_logo, x_position, y_position, label, value)
    
    for label, value in fields_mode:
        if mode == 'Cash':
            exit
        elif mode == 'Cheque':
            if label in ['Cheque Date:','Cheque No:','IFSC:', 'Account No:']:
                y_position = print_label_value(c, x_logo, x_position, y_position, label, value)
        elif mode == 'EFT':
            if label in ['EFT Date:', 'UTR No:', 'IFSC:', 'Bank Name:', 'Bank Branch:']:
                y_position = print_label_value(c, x_logo, x_position, y_position, label, value)
