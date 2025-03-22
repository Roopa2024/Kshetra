import tkinter as tk
from tkinter import messagebox
import pandas as pd
import openpyxl
from openpyxl import load_workbook
import os, configparser
from openpyxl import Workbook
import barcode
from barcode.writer import ImageWriter
import pdf_data
from PIL import Image, ImageFont, ImageDraw

config_path = os.path.abspath(os.path.join(os.path.dirname(__file__), "config", "receipt.ini"))
custom_font_path = os.path.dirname(os.path.abspath(__file__))+"\config\RobotoMono-Regular.ttf"
print(config_path)
config = configparser.ConfigParser()
config.read(config_path)
xcl_file = config.get('Filenames', 'xcl_file')
xcl_sheet = config.get('Filenames', 'xcl_sheet')
pdf_heading = config['Heading']['pdf_heading']
pdf_headings = pdf_heading.split(',')
try:
    font = ImageFont.truetype(custom_font_path, 20)
except OSError:
    print(f"Font {custom_font_path} not found, using default.")
    font = ImageFont.load_default()
    
def increment_counter(col_number):
    if not os.path.exists(xcl_file):
        print(f"⚠️ File '{xcl_file}' not found. Creating a new one...{col_number}")
        wb = Workbook()  # Create a new workbook
        wb.active.append(["Counter"])  # Add a header row (optional)
        wb.save(xcl_file)  # Save new workbook
        print("✅ New file created successfully!")

    print(f"increment_counter is col # = {col_number}")
    wb = load_workbook(xcl_file)   # Load the workbook and select the sheet
    sheet = wb.active               # Or use wb['SheetName'] if you have a specific sheet
    last_row = sheet.max_row        # Get the last row number
    last_value = sheet.cell(row=last_row, column=col_number).value  # Adjust column index as needed
    print(f"Value is {last_value}")

    if(col_number == 2):
        #match = re.match(r'([a-zA-Z]+)(\d+)', last_value)
        match = last_value[-4:] 

        if match:
            prefix = remaining_text = last_value[:-4]    # Extract the prefix and the number part
            num_part = last_value[-4:] 
        
            # Increment the numeric part and pad it to 4 digits (e.g., 0002)
            num_part = str(int(num_part) + 1).zfill(len(num_part))  # Use zfill to maintain leading zeros
        
            new_value = prefix + num_part # Reassemble the string with the incremented value
        else:
            print("Invalid format")
            return None    
    else:
        new_value = last_value + 1

    print(f"New value: {new_value}")
    return new_value
    
def get_bar_directory(pdf_filename):
    base_dir = os.getcwd()
    pdf_name_without_ext = os.path.splitext(pdf_filename)[0]
    output_folder = os.path.join(base_dir, "barcodes", pdf_name_without_ext)
    os.makedirs(output_folder, exist_ok=True)
    return output_folder

def save_to_excel(**kwargs):
    for key, value in kwargs.items():
        print(f"{key}: {value}")
    data = {key: [value] for key, value in kwargs.items()} # Convert each value in kwargs to a list
      
    df = pd.DataFrame(data) # Convert dictionary to pandas DataFrame
    print (f"Data is {data}")
    
    # If the file exists, append the new data to it, else create a new file
    try:
        existing_df = pd.read_excel(xcl_file)         # Check if file exists
        new_df = pd.concat([existing_df, df], ignore_index=True)
        new_df.to_excel(xcl_file, index=False)
    except FileNotFoundError:
        df.to_excel(xcl_file, index=False)         # Create a new Excel file if it does not exist

    #messagebox.showinfo("Success", "Data saved successfully to Excel!")

def draw_text(barcode_path, text_label):
    barcode_image = Image.open(barcode_path)                # Open the barcode image and get dimensions
    img_width, img_height = barcode_image.size
    new_height = img_height + 40                            # Create a new image with extra space ABOVE the barcode
    new_image = Image.new("RGB", (img_width, new_height), "white")
    draw = ImageDraw.Draw(new_image)
    bbox = draw.textbbox((0, 0), text_label, font=font)     # Get text bounding box
    text_width = bbox[2] - bbox[0]                          # Calculate width
    text_height = bbox[3] - bbox[1] 
    text_x = img_width - (text_width + 35)
    text_y = 20                                             # Place text at the top
    draw.text((text_x, text_y), text_label, fill="black", font=font)
    new_image.paste(barcode_image, (0, 45))                 # Paste barcode BELOW the text
    new_image.save(barcode_path)
  
def generate_barcode():
    try:
        df = pd.read_excel(xcl_file, sheet_name=xcl_sheet)
        workbook = openpyxl.load_workbook(xcl_file)
        sheet_name = workbook.active
        file_names = df.iloc[:, 0]
        bar_dir = get_bar_directory(pdf_headings[0]) #Selects the trust name, hardcoded for now
        os.makedirs(bar_dir, exist_ok=True)
    except ValueError:
        print("Error: Sheet not found!")
        messagebox.showinfo("Error", f"Sheet '{sheet_name}' not found in the file.")
        return

    column_name = df.columns[1]                             #barcodes are stored in the first column
    for index, row in df.iterrows():
        barcode_data = str(row[column_name])                # Convert to string
        text_label = str(row['TextColumn'])
        code128 = barcode.get_barcode_class('code128')      # Select barcode type
        barcode_obj = code128(barcode_data, writer=ImageWriter())
        options = {                                         # Define barcode settings
            "module_height": 7,                             # ⬅ Reduce barcode height (default: ~50)
            "font_size": 10,                                # Adjust text size below barcode
            "font_path": custom_font_path                   #"C:/Users/RoopaHegde/Downloads/RobotoMono-Regular.ttf",
        }
        file_name = os.path.join(bar_dir, f"{index}")
        file_name_png = os.path.join(bar_dir, f"{index}.png")
        barcode_obj.save(file_name, options=options) 
        draw_text(file_name_png, text_label)
        #draw_text(barcode_obj, text_label)

def load_column_values_to_dropdown(column_name):
    df = pd.read_excel(xcl_file, xcl_sheet)      # Read the spreadsheet and load the column values
    if column_name not in df.columns:     # Ensure the column exists in the DataFrame
        raise ValueError(f"Column '{column_name}' not found in the sheet '{xcl_sheet}'")
        
    dropdown_values = df[column_name].dropna().tolist()  # Drop any NaN values and convert to list
    return dropdown_values
