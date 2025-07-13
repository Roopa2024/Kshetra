from tkinter import messagebox
import pandas as pd
from openpyxl import load_workbook, Workbook
from openpyxl.drawing.image import Image
from openpyxl.styles import PatternFill
from openpyxl.styles import Font
import os, configparser, io
import barcode, qrcode
from barcode.writer import ImageWriter
from PIL import Image as PILImage, ImageFont, ImageDraw
from datetime import datetime
import qrcode
import sys, hmac
from Crypto.Hash import SHA3_256

config_path = os.path.abspath(os.path.join(os.path.dirname(__file__), "config", "receipt.ini"))
custom_font_path = os.path.dirname(os.path.abspath(__file__))+r"\config\RobotoMono-Regular.ttf"
config = configparser.ConfigParser()
config.read(config_path)
xcl_file = config.get('Filenames', 'xcl_file')
xcl_sheet = config.get('Filenames', 'xcl_sheet')
secret = config.get('Filenames', 'secret_key')
secret_key = secret.split(',')
pdf_heading = config['Heading']['pdf_heading']
pdf_headings = pdf_heading.split(',')
receipt_entity_xcl = config.get('Filenames', 'receipt_input_files')
receipt_entity_xcls = receipt_entity_xcl.split(',')
receipt_mapping_xcl = config.get('Filenames', 'receipt_mapping_files')
receipt_mapping_xcls = receipt_mapping_xcl.split(',')

code_mappings = config.get('Filenames', 'code_mapping_sheet')

try:
    font = ImageFont.truetype(custom_font_path, 20)
except OSError:
    print(f"Font {custom_font_path} not found, using default.")
    font = ImageFont.load_default()

def hmac_sha3(key, message):
    key = key.encode()
    message = message.encode()
    hmac_hash = hmac.new(key, message, SHA3_256)
    #print(hmac_hash.hexdigest())
    first_5 = hmac_hash.hexdigest()[:5]
    return first_5

#Function to increment the serial number in excel
def increment_counter(index, xcl_path):
    if not os.path.exists(xcl_path):
        print(f"File '{xcl_path}' not found. Creating a new one...")
        wb = Workbook()                                             # Create a new workbook
        wb.active.append(["Counter"])                               # Add a header row (optional)
        wb.save(xcl_path)                                           # Save new workbook
        print("New file created successfully!")

    df = pd.read_excel(xcl_path, sheet_name=xcl_sheet, header=1) 
    wb = load_workbook(xcl_path)                                    # Load the workbook and select the sheet
    last_id = df["Id."].dropna().iloc[-1] 
    new_id = int(last_id) + 1

    last_globeid= df["Globe ID"].dropna().iloc[-1]

    if last_globeid:
        prefix = last_globeid[:-4]               # Extract the prefix and the number part
        num_part = last_globeid[-4:]                              # Increment the numeric part and pad it to 4 digits (e.g., 0002)
        num_part = str(int(num_part) + 1).zfill(len(num_part))  # Use zfill to maintain leading zeros
        new_globeid = prefix + num_part                           # Reassemble the string with the incremented value
        #print(f"NUM part {num_part} and new value {new_globeid}")

        if new_globeid:
            textcol = hmac_sha3(secret_key[index], new_globeid) 
        return new_id, new_globeid, textcol
    else:
        print("Globe ID not found")
        return None    

def get_trans_type(mapping_path, mode):
    print(f"MODE is {mode} EXcel is {mapping_path}")

    if not os.path.exists(mapping_path):
        print(f"File '{mapping_path}' not found.")
        return
    df = pd.read_excel(mapping_path, sheet_name=code_mappings, header=0) 
    mask = df.eq(mode)
    if not mask.values.any():                                               # Ensure we found at least one match
        raise ValueError(f"No cell found with value {mode}")
    row_idx, col_idx = divmod(mask.values.argmax(), mask.shape[1])          # Get first match's row, col indices
    value = df.iat[row_idx, col_idx + 1]                                    # Get value from the next column                              

    print(value)
    return value

# Save data to the DB(excel sheet here)
def save_to_excel(file_path, id, selected_entity, **kwargs):
    print(f"save_to_excel : {file_path}")
    filename = os.path.basename(file_path)
    try:
        wb = load_workbook(file_path)                                   # Load the existing Excel file or create a new one
        ws = wb[xcl_sheet]  #wb.active
    except FileNotFoundError:
        from openpyxl import Workbook
        wb = Workbook()
        ws = wb[xcl_sheet]  #wb.active
        ws.append(list(kwargs.keys()))                                  

    next_row = id + 2    

    if 'receipt' in file_path:
        code_folder = 'Receipt'
    elif 'voucher' in file_path:
        code_folder = 'Voucher'
    elif 'invoice' in file_path:
        code_folder = 'Invoice'

    save_qrcode_path = (f"data/code/{selected_entity}/{code_folder}/qrcode/{id}.png")
    save_qrcode_path = os.path.abspath(save_qrcode_path)                                              # Determine the next available row
    qr_code_path = generate_qr_code(kwargs, save_qrcode_path) #, 'tmp_qr.png')                             # Generate QR Code before inserting data

    # Insert data into the correct columns
    for col_num, (key, value) in enumerate(kwargs.items(), start=2):
        if key not in ["Bar Code", "QR Code"]:   #to retain the existing values of these cols in excel
            #print(f"Key = {key} col = {col_num} and value = {value}")
            ws.cell(row=next_row, column=col_num, value=value)

    # Insert barcode into the "Bar Code" column (if exists)
    # if "Bar Code" in kwargs and kwargs["Bar Code"]:
    #     barcode_img = kwargs["Bar Code"]  
    #     img = Image(barcode_img)
    #     barcode_col = list(kwargs.keys()).index("Bar Code") + 1          # Find the column index for "Bar Code"
    #     ws.add_image(img, ws.cell(row=next_row, column=barcode_col).coordinate)

    # # Insert QR Code image
    # if "QR Code" in kwargs:
    #     img = Image(qr_code_path)
    #     qr_col = list(kwargs.keys()).index("QR Code") + 1
    #     ws.add_image(img, ws.cell(row=next_row, column=qr_col).coordinate)

    wb.save(file_path)                                                  # Save the workbook

# Function to handle bracode and TextColumn placement
def draw_text(barcode_path, text_label, save_barcode_path):
    barcode_image = PILImage.open(barcode_path)                         # Open the barcode image and get dimensions
    img_width, img_height = barcode_image.size
    new_height = img_height + 40                                        # Create a new image with extra space ABOVE the barcode
    new_image = PILImage.new("RGB", (img_width, new_height), "white")
    draw = ImageDraw.Draw(new_image)
    bbox = draw.textbbox((0, 0), text_label, font=font)                 # Get text bounding box
    text_width = bbox[2] - bbox[0]                                      # Calculate width
    text_height = bbox[3] - bbox[1] 
    text_x = img_width - (text_width + 35)
    text_y = 20                                                         # Place text at the top
    draw.text((text_x, text_y), text_label, fill="black", font=font)
    new_image.paste(barcode_image, (0, 45))                             # Paste barcode BELOW the text
    new_image.save(barcode_path)
    if not os.path.exists(save_barcode_path):
        os.makedirs(os.path.dirname(save_barcode_path), exist_ok=True)
    new_image.save(save_barcode_path)

# Function to generate barcode
def generate_barcode(data, temp_path="temp_barcode"):
    data = str(data) 
    barcode_class = barcode.get_barcode_class("code128")
    barcode_obj = barcode_class(data, writer=ImageWriter())
    options = {                                                         # Define barcode settings
            "module_height": 7,                                         # ⬅ Reduce barcode height (default: ~50)
            "font_size": 10,                                            # Adjust text size below barcode
            "font_path": custom_font_path                               #"C:/Users/RoopaHegde/Downloads/RobotoMono-Regular.ttf",
        }
    barcode_obj.save(temp_path, options)
    return temp_path

# Function to generate QR code
def generate_qr_code(row_data, save_qrcode_path, qr_path="temp_qr.png"):
    # Create a copy of row_data to prevent modifying the original dictionary
    qr_data = {key: value for key, value in row_data.items() if key != "Bar Code" and value}

    # Convert row data to a readable string (":" and "-" are not readble by QR Code)
    qr_content = ";".join([f"{key.replace(':', ' ')} = {str(value).replace(':', ' ')}" for key, value in qr_data.items()])
    #print("QR Code Content:\n", qr_content)
    
    # Generate QR Code with error correction and force UTF-8 encoding
    qr = qrcode.QRCode(
        version=6,                                                      # Controls the size of the QR Code (1 is the smallest)
        error_correction=qrcode.constants.ERROR_CORRECT_L,              # Allows for small errors in reading
        box_size=10,                                                    # Size of each box in the QR Code
        border=2,                                                       # Border around the QR Code
    )
    qr.add_data(qr_content)
    qr.make(fit=True)
    img = qr.make_image(fill="black", back_color="white")               # Create an image from the QR Code
    img.save(qr_path)  
    if not os.path.exists(save_qrcode_path):
        os.makedirs(os.path.dirname(save_qrcode_path), exist_ok=True)
    img.save(save_qrcode_path)                                                 # Save QR Code
    return qr_path

def cancel_last_row(excel_path):
    wb = load_workbook(excel_path)
    ws = wb[xcl_sheet] #wb.active

    header = [cell.value for cell in ws[2]]                             # Find the header row and column index of "Globe stat"
    try:
        col_index = header.index("Globe Stat") + 1                     # openpyxl is 1-indexed
    except ValueError:
        raise Exception("Column 'Globe stat' not found")

    # Get the last row and update the cell
    last_row = ws.max_row
    ws.cell(row=last_row, column=col_index, value="cancelled")

    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    for cell in ws[ws.max_row]:                         # Highlight each cell in the last row
        cell.fill = yellow_fill

    strike_font = Font(strike=True)
    for cell in ws[ws.max_row]:                         # Apply strikethrough to each cell in the last row
        cell.font = strike_font
    # Save changes
    wb.save(excel_path)
    messagebox.showinfo(title="Success", message=f"Row {last_row} successfully cancelled")
