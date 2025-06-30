import tkinter as tk
from tkinter import messagebox
import pandas as pd
from openpyxl import load_workbook
import re, sys, os

file_path = "Images/cheque_data.xlsx"

def resource_path(rel_path):
    base = getattr(sys, "_MEIPASS", os.path.abspath(os.path.dirname(__file__)))
    return os.path.join(base, rel_path)

abs_path = resource_path(file_path)

def increment_counter():
    wb = load_workbook(abs_path)   # Load the workbook and select the sheet
    sheet = wb.active               # Or use wb['SheetName'] if you have a specific sheet
    last_row = sheet.max_row        # Get the last row number
    last_value = sheet.cell(row=last_row, column=1).value  # Adjust column index as needed
    print(f"Value is {last_value}")
    match = re.match(r'([a-zA-Z]+)(\d+)', last_value)
    
    if match:
        prefix = match.group(1)     # Extract the prefix and the number part
        num_part = match.group(2)
        
        # Increment the numeric part and pad it to 4 digits (e.g., 0002)
        num_part = str(int(num_part) + 1).zfill(len(num_part))  # Use zfill to maintain leading zeros
        
        new_value = prefix + num_part # Reassemble the string with the incremented value
        print(f"New value: {new_value}")
        return new_value
    else:
        print("Invalid format")
        return None

def save_to_excel(**kwargs):
    for key, value in kwargs.items():
        print(f"{key}: {value}")
    # Convert each value in kwargs to a list
    data = {key: [value] for key, value in kwargs.items()}
      
    # Convert dictionary to pandas DataFrame
    df = pd.DataFrame(data)
    
    # If the file exists, append the new data to it, else create a new file
    try:
        # Check if file exists
        existing_df = pd.read_excel(abs_path)
        new_df = pd.concat([existing_df, df], ignore_index=True)
        new_df.to_excel(abs_path, index=False)
    except FileNotFoundError:
        # Create a new Excel file if it does not exist
        df.to_excel(abs_path, index=False)

    # Show success message
    messagebox.showinfo("Success", "Data saved successfully to Excel!")
