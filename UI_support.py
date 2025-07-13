import tkinter as tk
import openpyxl, os, sys, shutil, configparser, getpass
import canvas_update, excel_data, pdf_data
import datetime
from datetime import date
import pandas as pd
from tkinter import messagebox
from tkinter import ttk
from tkcalendar import DateEntry

# Load configurations
config_path = os.path.join(os.path.dirname(__file__), "config", "receipt.ini")
config = configparser.ConfigParser()
config.read(config_path)

font_settings = (config.get('FontSettings', 'font_name'), config.getint('FontSettings', 'font_size'))
entity_name = config.get('Filenames', 'entity_name')
entity_names = entity_name.split(',')
receipt_entity_xcl = config.get('Filenames', 'receipt_input_files')
receipt_entity_xcls = receipt_entity_xcl.split(',')
receipt_mapping_xcl = config.get('Filenames', 'receipt_mapping_files')
receipt_mapping_xcls = receipt_mapping_xcl.split(',')
voucher_entity_xcl = config.get('Filenames', 'voucher_input_files')
voucher_entity_xcls = voucher_entity_xcl.split(',')
voucher_mapping_xcl = config.get('Filenames', 'voucher_mapping_files')
voucher_mapping_xcls = voucher_mapping_xcl.split(',')
invoice_entity_xcl = config.get('Filenames', 'invoice_input_files')
invoice_entity_xcls = invoice_entity_xcl.split(',')
invoice_mapping_xcl = config.get('Filenames', 'invoice_mapping_files')
invoice_mapping_xcls = invoice_mapping_xcl.split(',')

#Mapping sheet name
mapping = config.get('Filenames', 'mapping_sheet')

width = config.getint('FontSettings', 'width')
pdf_heading = config['Heading']['pdf_heading']
pdf_headings = pdf_heading.split(',')
entity = pdf_heading.split(',')
voucher_pdf_heading = config['Heading']['voucher_pdf_heading']
#voucher_pdf_headings = voucher_pdf_heading.split(',')
voucher_entity = voucher_pdf_heading.split(',')
invoice_pdf_heading = config['Heading']['invoice_pdf_heading']
invoice_entity = invoice_pdf_heading.split(',')
copy_type = config['Heading']['copy_type']
copy_types = copy_type.split(',')
xcl_sheet = config.get('Filenames', 'xcl_sheet')
printer_name = config.get('printer', 'printer')

# Function to limit input entry length in UI
def limit_chars(var, limit):
    if len(var.get()) > limit:
        var.set(var.get()[:limit])

# Function to limit characters in a Text widget (e.g., Address)
def limit_text_chars(event, text_widget, limit):
    current_text = text_widget.get("1.0", "end-1c")  # Get text without newline at the end
    if len(current_text) > limit:
        text_widget.delete("1.0", "end")  # Clear text
        text_widget.insert("1.0", current_text[:limit])

# Function to get base path for EXE
def resource_path(relative_path):
    """ Get absolute path to resource, works for dev and for PyInstaller """
    try:
        # PyInstaller creates a temp folder and stores path in _MEIPASS
        base_path = sys._MEIPASS
    except AttributeError:
        base_path = os.path.abspath(".")
    return os.path.join(base_path, relative_path)

def validate_alpha(new_char, current_value):
    if not new_char.isalpha() and new_char != " ":
        return False
    return len(current_value) < 35

def validate_alpha_branch(new_char, current_value):
    if not new_char.isalpha() and new_char != " ":
        return False
    return len(current_value) < 20

def update_amount_in_words(entry_widget, text_widget):
    try:
        amount = float(entry_widget.get())
        from num2words import num2words
        amount_words = canvas_update.convert_to_words(amount) #num2words(amount, to="currency", lang="en_IN").title()
    except:
        amount_words = ""

    text_widget.configure(state="normal")
    text_widget.delete("1.0", tk.END)
    text_widget.insert(tk.END, amount_words)
    text_widget.configure(state="disabled")

def get_excel_file(selected_idx):
    print(f"get_excel_file {receipt_entity_xcls[selected_idx]}")
    excel_path = resource_path(receipt_entity_xcls[selected_idx])
    excel_mapping_path = resource_path(receipt_mapping_xcls[selected_idx])

    file_name = os.path.basename(excel_path)
    dest_excel_path = f"data/excel/{entity_names[selected_idx]}/{file_name}"
    print(f"DEST to copy excel {dest_excel_path}")

    if not os.path.exists(dest_excel_path):
        os.makedirs(os.path.dirname(dest_excel_path), exist_ok=True)
        shutil.copy(excel_path, dest_excel_path)
        print(f"Copied {excel_path} to {dest_excel_path}")
    else:
        print(f"Excel File already exists. {dest_excel_path}")
    return dest_excel_path,excel_mapping_path

# Function to get dropdown values from the Excel sheet
def get_dropdown_values(file_path, sheet_name, column_name):
    workbook = openpyxl.load_workbook(file_path, data_only=True)  # Load Excel
    sheet = workbook[sheet_name]  # Select sheet

     # Find column index by header name
    headers = [cell.value for cell in sheet[1]]  # Read header row (1st row)
    if column_name not in headers:
        print(f"Error: Column '{column_name}' not found in sheet '{sheet_name}'")
        return []
    
    col_index = headers.index(column_name) + 1  # Convert to 1-based index

    # Extract values from the column (starting from row 2)
    values = [
        sheet.cell(row=i, column=col_index).value
        for i in range(2, sheet.max_row + 1)
        if sheet.cell(row=i, column=col_index).value is not None
    ]
    return values

def UI_amount_in_words(frame, entry, row_offset, width):
    entry.bind("<KeyRelease>", lambda event, e=entry: update_amount_in_words(e, amount_in_words_text))
    # Add amount in words field below
    #row_offset = len(fields)
    tk.Label(frame, text="Amount in Words:", font=font_settings, bg="#99ccff").grid(row=row_offset, column=0, sticky="w", pady=5)
    amount_in_words_text = tk.Text(frame, font=font_settings, height=2, wrap="word", width=width)
    amount_in_words_text.grid(row=row_offset, column=1, padx=5, pady=5, sticky="ew")
    amount_in_words_text.configure(state="disabled")
    return amount_in_words_text

input_vars = {}                                                                     # Store variables for later access
# Function to add widgets dynamically to each group
def add_widgets_to_group(frame, fields):
    amount_in_words_var = tk.StringVar()

    for i, (label, widget, field_options) in enumerate(fields):
        #For the heading "Contribution"
        if label == "__HEADER__":
            tk.Label(frame, text=widget, font=field_options.get("font", font_settings), bg="#99ccff").grid(
                row=i, column=0, columnspan=2, pady=10, sticky="w"
            )
            continue

        tk.Label(frame, text=label, font=font_settings, bg="#99ccff").grid(row=i, column=0, sticky="w", pady=5)
        
        var = field_options.get("textvariable", None)               # Get the variable, if any, from field_options
        if var:
            limit = 10 if "PAN:" in label else 27 if "UTRN:" in label else 100
            var.trace_add("write", lambda *args, v=var, l=limit: limit_chars(v, l))
            input_vars[label] = var

        # Handle OptionMenu separately, since it requires both 'variable' and 'value' parameters
        if widget == tk.OptionMenu:
            var = field_options.get("variable")                                     # Ensure the 'variable' is passed
            options = field_options.get("options")                                  # Ensure the options are passed
            if var is not None and options is not None:
                entry = widget(frame, var, *options)
            else:
                raise ValueError(f"OptionMenu requires both 'variable' and 'options' in field_options. Missing for '{label}'")
        elif label == 'Receipt Date:':
            entry = widget(frame, font=font_settings, width=25, **field_options, state="readonly")
        elif widget == tk.Text:
            entry = widget(frame, font=font_settings, height=field_options.get("height"), width=40)
            entry.config(wrap="word")  # Set wrap here
            entry.bind("<KeyRelease>", lambda event, e=entry: limit_text_chars(event, e, 200))
        else:
            entry = widget(frame, font=font_settings, width=width, **field_options)

        entry.grid(row=i, column=1, padx=5, pady=(0,5), sticky="ew")                            # Add the widget to the grid
        input_vars[label] = entry

        if label == "Amount:":
            amount_in_words_text = UI_amount_in_words(frame, entry, len(fields), 60)
            input_vars["Amount in Words:"] = amount_in_words_text

def generate_excel_file(selected_entity, selected_index):
    print(f"generate_excel_file {selected_entity}")
    entity_path = resource_path(selected_entity)
    file_name = os.path.basename(entity_path)
    dest_excel_path = f"data/excel/{entity_names[selected_index]}/{file_name}"
    if not os.path.exists(dest_excel_path):
        os.makedirs(os.path.dirname(dest_excel_path), exist_ok=True)
        shutil.copy(entity_path, dest_excel_path)
        print(f"Copied new {entity_path} to {dest_excel_path}")
    else:
        print(f"File {dest_excel_path} already exists.")
    return dest_excel_path

def get_cheque_data(voucher):
    cheque_data = voucher.get('cheque', {})
    cheque_date = cheque_data.get('Cheque Date:')
    cheque_no = cheque_data.get("Cheque No.:")
    cheque_IFSC = cheque_data.get('IFSC Code:')
    cheque_AC_no = cheque_data.get('Account No.:')
            
    if cheque_no:
        cheque_no_val = cheque_no.get()
    else:
        cheque_no_val = ''
    return cheque_date.get(), cheque_no_val, cheque_IFSC.get(), cheque_AC_no.get()

def get_eft_data(voucher):
    utrn_val, bank, branch = '', '', ''
    eft_data = voucher.get('eft', {})
    eft_date = eft_data.get('Bank Date:')
    utrn = eft_data.get("UTRN:")
    eft_IFSC = eft_data.get('IFSC Code:')
    bank_name = eft_data.get('Bank Name:')
    bank_branch = eft_data.get('Bank Branch:')

    if utrn:
        utrn_val = utrn.get()
        ifsc = eft_IFSC.get()
        bank = bank_name.get()
        branch = bank_branch.get()
    
    return eft_date.get(), utrn_val, ifsc, bank, branch


def delete_last_n_rows(file_path, n, sheet_name):
    wb = openpyxl.load_workbook(file_path)
    ws = wb[sheet_name] if sheet_name else wb.active

    max_row = ws.max_row
    start_row = max_row - n + 1  # Row to start deleting from

    if start_row >= 1:
        print(f"delete rows {start_row}")
        ws.delete_rows(start_row, n)
        wb.save(file_path)
    else:
        print("Not enough rows to delete.")

def clear_voucher_fields(voucher):
    for key, widget in voucher.items():
        try:
            if isinstance(widget, dict):  # Check if the widget is a dictionary
                for sub_key, sub_widget in widget.items():
                    if isinstance(sub_widget, DateEntry):
                        sub_widget.set_date(date.today())  # Clear the date
                    elif isinstance(sub_widget, tk.Entry):
                        sub_widget.delete(0, tk.END)
                    elif isinstance(sub_widget, ttk.Combobox):
                        sub_widget.set('')
                    elif hasattr(sub_widget, 'set'):
                        sub_widget.set('')
            elif isinstance(widget, DateEntry):
                widget.set_date(date.today())  # or use widget.delete(0, tk.END) if needed
            elif isinstance(widget, tk.Entry):
                widget.delete(0, tk.END)
            elif isinstance(widget, ttk.Combobox):
                widget.set('')
            elif hasattr(widget, 'set'):
                widget.set('')
            elif isinstance(widget, tk.Text):
                current_state = widget.cget("state")
                if current_state == "disabled":
                    widget.configure(state="normal")
                    widget.delete("1.0", tk.END)
                    widget.configure(state="disabled")
                else:
                    widget.delete("1.0", tk.END)

            # Clear readonly Entry bound to StringVar (e.g., 'exp_type')
            expense_entry = voucher['exp_type']
            expense_var = expense_entry.cget("textvariable")
            var_obj = expense_entry.getvar(expense_var)
            # Temporarily set to normal, clear, then restore
            expense_entry.config(state="normal")
            expense_entry.setvar(expense_var, '')
            expense_entry.config(state="readonly")

            if isinstance(voucher['payment_mode'], tk.StringVar):
                voucher['payment_mode'].set('Cash') # = 'Cash'  # Unselect all
        except Exception as e:
            print(f"Error clearing field {key}: {e}")

def get_user(amount, path):
    if amount:
        username = getpass.getuser()
        if 'invoice' in path:
            authoriser = f'Printed By: {username}'
        else:
            authoriser = f'Issued By: {username}'
        return authoriser
    else:
        authoriser = ''
        return

def voucher_print(path, mapping_voucher_path, voucher_entries, selected_indx, checkbox_var, selected_option, cheque_frame, online_frame):
    dest_excel_path = generate_excel_file(path, selected_indx) 
    print(f"DEST EXCEL {dest_excel_path}")
    if 'invoice' in dest_excel_path:
        mode = selected_option.get()
    else:
        for index, var in selected_option.items():
            selected_option[index].set(var.get())
            print(f"1. VAR = {var} and selection is set to {selected_option[index]}")

    for i, voucher in enumerate(voucher_entries):
        #cheque_date, cheque_no_val, cheque_IFSC, cheque_AC_no, eft_date, utrn_val, bank_name, bank_branch = '', '', '', '','','','',''
        inst_date, inst_no_val, IFSC, inst_AC_no, bank_name, bank_branch = '', '', '', '','','',
        inst_voucher_date, inst_voucher_num = '', ''
        pan_entry = voucher.get('pan')
        if hasattr(pan_entry, 'get'):
            pan = pan_entry.get()
        else:
            pan = ''
        inv_entry = voucher.get('inv_no')
        if hasattr(inv_entry, 'get'):
            inv_no = inv_entry.get()
        else:
            inv_no = ''

        voucher_date = voucher['date'].get()
        pay_to = voucher['pay_to'].get()
        amount = voucher['amount'].get()
        #pan = voucher['pan'].get()
        if 'invoice' in path:
            addr_text = voucher['addr']  # Your Text widget
            addr = addr_text.get("1.0", "end-1c")
            challan = voucher['challan'].get()
            cert = voucher['cert'].get()
            CIN = challan + ' ' + cert
            print(f"CIN is {CIN} MODE is {mode}")
            #mode = voucher['payment_mode']
        else:
            addr = voucher['addr'].get()
            CIN = ''
            mode = voucher['payment_mode'].get()
            #mode = selected_option[i].get()

        pur_code = voucher['pur_code'].get()
        pur_head = voucher['pur_head'].get()
        pur_cat = voucher['pur_cat'].get()
        exp_type = voucher['exp_type'].get()

        if not amount:
            user_choice = messagebox.askyesno("Amount Missing","Do you want to continue without entering an amount?")
            if user_choice:
                amount = ""  # User wants to continue; you can handle this case as needed
            else:
                print(f"INDEX = {i}")
                delete_last_n_rows(dest_excel_path, n=i, sheet_name=xcl_sheet)
                return
        if mode == "Cheque":
            cheque_date, cheque_no_val, cheque_IFSC, cheque_AC_no = get_cheque_data(voucher)
            inst_voucher_date = cheque_date
            inst_voucher_num = cheque_no_val
            IFSC = cheque_IFSC
        elif mode == "EFT":
            eft_date, utrn_val, eft_ifsc, bank_name, bank_branch = get_eft_data(voucher)
            inst_voucher_date = eft_date
            inst_voucher_num = utrn_val
            IFSC = eft_ifsc

        id, globe_id, bar_text = excel_data.increment_counter(selected_indx, dest_excel_path) 
        date = datetime.datetime.now()
        formatted_date = date.strftime("%d/%m/%Y %H.%M.%S")

        if 'invoice' in dest_excel_path:
            voucher_no = inv_no
            code_folder = 'Invoice'
        else:
            voucher_no = globe_id
            code_folder = 'Voucher'
        
        barcode_path = excel_data.generate_barcode(str(globe_id))   #barcode is temporarily saved in the current dir
        save_barcode_path = (f"data/code/{entity_names[selected_indx]}/{code_folder}/barcode/{id}.png")
        save_barcode_path = os.path.abspath(save_barcode_path)
        print(f"VOUCHER BARCODE PATH {save_barcode_path}")
        excel_data.draw_text(f"{barcode_path}.png", bar_text, save_barcode_path) 

        print(f"VOUCHER Mode = {mode}")
        trans_type = excel_data.get_trans_type(mapping_voucher_path, mode)

        #the col header and order here should be an exact match with the excel.
        kwargs = {
            'Id.' : id,
            'Payee Name' : pay_to,
            'Address': addr,
            'Amount' : amount,
            'Voucher No.' : voucher_no, 
            'Voucher Date' : voucher_date,
            'Payment Mode': mode,
            'Transaction Type': trans_type,
            'Bank Realisation Date' : '',
            'Inst. Date': inst_voucher_date,
            'Inst. No': inst_voucher_num,
            'IFSC': IFSC,
            'Account No': inst_AC_no,
            'Bank Name': bank_name,
            'Bank Branch': bank_branch,
            'Payee PAN No' : pan,
            'Payee GST No' : '',
            'GST Stat' : '',
            'GST CPIN' : '',
            'TDS Stat' : '',
            'TDS CIN' : CIN,
            'Expense ID' : '',
            'Globe ID' : globe_id,
            'Globe Stat' : '',
            'Expense Type':exp_type,
            'Purchase Code': pur_code,
            'Purchase Head': pur_head,
            'Purchase Category': pur_cat,
            'User Description' : '',
            'Proof' : '',
            'Narration' : '',
            'Print Date' : formatted_date,
            'Issuer' : get_user(amount, path),
            'Bar Code' : f"{barcode_path}.png",
            'TextColumn' : bar_text,
            'QR Code' : '',
        }

        #print(kwargs)
        excel_data.save_to_excel(dest_excel_path, id, entity_names[selected_indx], **kwargs)
    
    if 'invoice' in path:
        selected_entity = "Invoice.pdf" #invoice_entity[selected_indx]
    else:
        selected_entity = "Voucher.pdf" #voucher_entity[selected_indx]

    folder_name = os.path.splitext(selected_entity)[0]       
    os.makedirs(f"data/pdfs/{entity_names[selected_indx]}/{folder_name}", exist_ok=True) 
    pdf_path = f"data/pdfs/{entity_names[selected_indx]}/{folder_name}/{kwargs['Id.']}.pdf"
    print(f"PDF PATH {pdf_path} entity_name = {entity_names[selected_indx]} and selected = {selected_entity}")
    
    pdf_data.create_pdf_from_kwargs_voucher(kwargs, voucher_entries, pdf_path, selected_entity, checkbox_var.get(), dest_excel_path, selected_indx, trans_type, mode)
    
    for i, voucher in enumerate(voucher_entries):
        clear_voucher_fields(voucher)
    
    if 'invoice' in path:
        selected_option.set('Cash')
        cheque_frame.grid() if selected_option == "Cheque" else cheque_frame.grid_remove()
        online_frame.grid() if selected_option == "EFT" else online_frame.grid_remove()
        print(f'RESET PAYMENT MODE  = {selected_option.get()}')
    else:
        for index, var in selected_option.items():
            print (f"RESET for voucher_{index} = {var.get()}")
            var.set("Cash")
            #selected_option[index] = tk.StringVar(value="Cash")
            #selected_value = selected_option[index] #.get()
            print(f"CLEAR VAR = {var.get()}")
        #         #selected_value = 'Cash'
            #voucher_entries['payment_mode'].set(selected_option[index])
            cheque_frame.grid() if var.get() == "Cheque" else cheque_frame.grid_remove()
            online_frame.grid() if var.get() == "EFT" else online_frame.grid_remove()

    cheque_frame.grid_remove()
    online_frame.grid_remove()

def clear_fields(widget):
    if isinstance(widget, DateEntry):
        widget.set_date(date.today())  
    elif isinstance(widget, tk.StringVar):
        widget.set("")
    elif isinstance(widget, ttk.Combobox):
        widget.set(widget["values"][0])
    elif isinstance(widget, tk.Entry):
        widget.delete(0, tk.END)
    elif isinstance(widget, tk.Text):
        current_state = widget.cget("state")
        if current_state == "disabled":
            widget.configure(state="normal")
            widget.delete("1.0", tk.END)
            widget.configure(state="disabled")
        else:
            widget.delete("1.0", tk.END)
    #checkbox_var.set(0)

#Receipt Cancel function
def cancel(selected_indx):
    dest_excel_path = generate_excel_file(receipt_entity_xcls[selected_indx], selected_indx)
    excel_data.cancel_last_row(dest_excel_path)


# Submit Function
def submit(selection_var, cheque_vars, online_vars, selected_indx, checkbox_var, cheque_frame, online_frame, mapping_receipt_path):   
    details = {}
    folder_name = 'Receipt'
    
    for label, var in input_vars.items():
        if isinstance(var, tk.StringVar):                       
            details[label] = var.get()
        elif isinstance(var, tk.Entry):                         # Normal Entry widget
            details[label] = var.get()
        elif isinstance(var, tk.Text):                          # Multiline Text widget
            details[label] = var.get("1.0", "end-1c").strip()   # Remove extra newlines
        else:
            details[label] = ""                                # Handle unexpected types safely
    #print("Details:", details)
    
    if selection_var.get() == "Cheque":
        cheque_details = {
            label: var.get() if isinstance(var, tk.StringVar) 
            else var.get("1.0", "end-1c") if isinstance(var, tk.Text)
            else var.get()
            for label, var in cheque_vars.items()}
        payment_mode = 'Cheque'
        cheque_date = cheque_details.get('Cheque Date:')
        cheque_no = cheque_details.get('Cheque No.:')
        IFSC = cheque_details.get('IFSC Code:')
        ac_no =  cheque_details.get('Account No.:')
        bank_name = cheque_details.get('Bank name:')
        bank_br = cheque_details.get('Branch:')
        bank_date, utrn = '', ''
    elif selection_var.get() in ("Cash", "EFT"):
        payment_mode = selection_var.get()
        cheque_date, cheque_no, IFSC, ac_no = '', '', '', ''
        if payment_mode == "EFT":
            online_details = {
                label: var.get() if isinstance(var, tk.StringVar) 
                else var.get("1.0", "end-1c") if isinstance(var, tk.Text)
                else var.get()
                for label, var in online_vars.items()}
            bank_date = online_details.get('Bank Date:')
            utrn = online_details.get('UTRN:')
            bank_name = online_details.get('Bank name:')
            bank_br = online_details.get('Branch:')
        elif payment_mode == 'Cash':
            bank_date, utrn, bank_name, bank_br = '', '', '',''

    if cheque_date:
        print(f'CHEQUE here {cheque_date} and {cheque_no}')
        inst_date = cheque_date
        inst_num = cheque_no
    else:
        print(f'CASh/EFT here {bank_date} and {utrn}')
        inst_date = bank_date
        inst_num = utrn

    dest_excel_path = generate_excel_file(receipt_entity_xcls[selected_indx], selected_indx)
    id, globe_id, bar_text = excel_data.increment_counter(selected_indx, dest_excel_path) 
    date = datetime.datetime.now()
    formatted_date = date.strftime("%d/%m/%Y %H.%M.%S")
    #authoriser = getpass.getuser()
    barcode_path = excel_data.generate_barcode(str(globe_id))   #barcode is temporarily saved in the current dir
    save_barcode_path = (f"data/code/{entity_names[selected_indx]}/Receipt/barcode/{id}.png")
    save_barcode_path = os.path.abspath(save_barcode_path)
    print(f"BARCODE PATH {save_barcode_path}")
    excel_data.draw_text(f"{barcode_path}.png", bar_text, save_barcode_path)       #adding TextColumn to the barcode

    trans_type = excel_data.get_trans_type(mapping_receipt_path, selection_var.get())

    #the col header and order here should be an exact match with the excel.
    kwargs = {
        'Id.' : id,
        'Receipt/Invoice Book No': '',
        'Receipt/Invoice No' : globe_id, 
        'Receipt/Invoice Date' : details.get("Receipt Date:"),
        'Amount' : details.get("Amount:"),
        'Contributor Name' : details.get("Contributor Name:"),
        'Contributor Address' : details.get("Address:"),
        'Contributor PAN' : details.get("PAN:"),
        'Contributor Type' : details.get("Contribution Type:"),
        'Contributor Intent' : details.get("Contribution Intent:"),
        'Payment Mode' : payment_mode,
        'Transaction Type' : trans_type,
        'Bank Realisation Date' : '',
        'Inst. Date' : inst_date,
        'Inst. No' : inst_num, 
        'IFSC' : IFSC,
        'Account No.' : ac_no,
        'Bank Name' : bank_name, 
        'Branch Name' : bank_br,
        'Income ID' : '',
        'Globe ID' : globe_id,
        'Globe Stat' : '',
        'Proof' : '',
        'Narration' : '',
        'Print Date' : formatted_date,
        'Issuer' : get_user(details.get("Amount:"), dest_excel_path),
        'Bar Code' : f"{barcode_path}.png",
        'TextColumn' : bar_text,
        'QR Code' : '',
    }

    excel_data.save_to_excel(dest_excel_path, id, entity_names[selected_indx], **kwargs)

    # save_qrcode_path = (f"data/code/{entity_names[selected_indx]}/Receipt/qrcode/{id}.png")
    # save_qrcode_path = os.path.abspath(save_qrcode_path)
    # qr_code_path = excel_data.generate_qr_code(kwargs, save_qrcode_path)
    #kwargs["QR Code"] = qr_code_path                                                           # Update kwargs with QR Code path
    
    entity_name = os.path.splitext(entity[selected_indx])[0]                        
    pdf_path = f"data/pdfs/{entity_name}/{folder_name}/{kwargs['Id.']}.pdf"
    os.makedirs(f"data/pdfs/{entity_name}/{folder_name}", exist_ok=True)

    with_bg = checkbox_var.get()
    if with_bg:
        pdf = 1
        pdf_data.create_pdf_from_kwargs(kwargs, pdf_path, entity_name, checkbox_var.get(), copy_types[1], copy_types[2], pdf, selected_indx)
        pdf_data.create_pdf_from_kwargs(kwargs, pdf_path, entity_name, checkbox_var.get(), copy_types[0], copy_types[0], pdf, selected_indx)
    else:
        pdf_data.create_pdf_from_kwargs(kwargs, pdf_path, entity_name, checkbox_var.get(), copy_types[1], copy_types[2], 0, selected_indx)
        pdf_data.create_pdf_from_kwargs(kwargs, pdf_path, entity_name, checkbox_var.get(), copy_types[1], copy_types[2], 1, selected_indx)   
        pdf_data.create_pdf_from_kwargs(kwargs, pdf_path, entity_name, checkbox_var.get(), copy_types[0], copy_types[0], 1, selected_indx)
    
    try:
        df = pd.read_excel(dest_excel_path, sheet_name=xcl_sheet, header=1)
        workbook = openpyxl.load_workbook(dest_excel_path)
        sheet = workbook[xcl_sheet] 
        pdf_dir = pdf_data.get_pdf_directory()
        os.makedirs(pdf_dir, exist_ok=True)
    except ValueError as e:
        print(f"Error: Worksheet named '{xcl_sheet}' not found. Please check the sheet name.")
        messagebox.showinfo(f"Error:", f"Worksheet named '{xcl_sheet}' not found. Please check the sheet name.") 
    except Exception as e:
        print(e)

    rows_data = []                                                                  # Loop through each row in the DataFrame and create a separate PDF
    for index, row in df.iterrows():
        row_dict = {df.columns[i]: row.iloc[i] for i in range(len(df.columns))}
        rows_data.insert(id, row_dict) 
        pdf_name = f"data/pdfs/{index}.pdf"
    workbook.save(dest_excel_path)

    for key, widget in cheque_vars.items():     
        clear_fields(widget)
    for key, widget in online_vars.items():
        clear_fields(widget)
    for key, widget in input_vars.items():
        clear_fields(widget)

    selection_var.set('Cash')
    cheque_frame.grid_remove()
    online_frame.grid_remove()
