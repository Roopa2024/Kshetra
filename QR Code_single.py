import pandas as pd
import qrcode
import os
from openpyxl import load_workbook
from openpyxl.drawing.image import Image
import io

# Read the Excel file into a DataFrame
excel_file = 'data1.xlsx'  # Replace with your file path
df = pd.read_excel(excel_file)

# Load the existing workbook
wb = load_workbook(excel_file)
ws = wb.active  # Or specify the sheet name: wb['Sheet1']

# Create a directory for QR codes if it doesn't exist
qr_code_dir = os.path.join(os.path.dirname(excel_file), 'qr_codes')
os.makedirs(qr_code_dir, exist_ok=True)

# Iterate over each row in the DataFrame
for index, row in df.iterrows():
    # Convert row values to a string for the QR code (you can customize this format)
    #f"{col}:
    row_data = ';'.join([f"{col}={row[col]}" for col in df.columns])

    image_width =100  # in pixels
    image_height = 100  # in pixels

    # Generate the QR code
    qr = qrcode.QRCode(
        version=10, 
        error_correction=qrcode.constants.ERROR_CORRECT_L, 
        box_size=10, 
        border=4
    )
    qr.add_data(row_data)
    qr.make(fit=True)
    print(row_data)
    # Save the QR code as an image to a BytesIO object (so we can insert it into Excel without saving it as a file)
    img_stream = io.BytesIO()
    img = qr.make_image(fill='black', back_color='white')
    img.save(img_stream)
    img_stream.seek(0)  # Move to the start of the stream

    # Create an Image object from the BytesIO stream
    img = Image(img_stream)
    img.width = image_width
    img.height = image_height

     # Set column width and row height (optional)
    ws.column_dimensions['K'].width = 20  # Adjust column B width
    for row in range(2, len(df) + 2):
        ws.row_dimensions[row].height = 100 
    ws['K1'] = 'QR Code' 
    # Insert the image into the Excel sheet (adjust the cell as needed, here inserting in column 'B')
    img.anchor = f'K{index + 2}'  # Adjust +2 to match the correct row in Excel (1-based index)
    
    # Add the image to the worksheet
    ws.add_image(img)

# Save the modified Excel file with QR codes
wb.save('qr_codes\data_with_qr_codes.xlsx')  # Save to a new file or overwrite the existing one

print("QR codes have been successfully added to the Excel file.")
