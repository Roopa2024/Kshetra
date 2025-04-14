import tkinter as tk
from tkinter import ttk
from tkinter import messagebox
from tkcalendar import DateEntry
import os, configparser, excel_data, pdf_data
import pandas as pd
import openpyxl
from openpyxl import load_workbook, Workbook
from openpyxl.drawing.image import Image
import datetime, getpass, sys, shutil
from PIL import Image, ImageTk

# Load configurations
config_path = os.path.join(os.path.dirname(__file__), "config", "receipt.ini")
config = configparser.ConfigParser()
config.read(config_path)
font_settings = (config.get('FontSettings', 'font_name'), config.getint('FontSettings', 'font_size'))
width = config.getint('FontSettings', 'width')
xcl_file = config.get('Filenames', 'xcl_file')
xcl_sheet = config.get('Filenames', 'xcl_sheet')
entity_xcl = config.get('Filenames', 'input_files')
mapping = config.get('Filenames', 'mapping_sheet')
entity_xcls = entity_xcl.split(',')
heading = config['Heading']['heading']
headings = heading.split(',')
pdf_heading = config['Heading']['pdf_heading']
pdf_headings = pdf_heading.split(',')
entity = pdf_heading.split(',')
copy_type = config['Heading']['copy_type']
copy_types = copy_type.split(',')

# Configuration to limit Access based on Login 
allowed_user = config['Access']['user']
allowed_users = allowed_user.split(',')
current_user = getpass.getuser()

print (f"User is {current_user}")

#if current_user not in allowed_users:
#    print(f"Access denied for user: {current_user}")
#    messagebox.showinfo(f"Error:", f"Access denied for user: {current_user}")
#    sys.exit(1)

# Function to get base path for EXE
def resource_path(relative_path):
    """ Get absolute path to resource, works for dev and for PyInstaller """
    try:
        # PyInstaller creates a temp folder and stores path in _MEIPASS
        base_path = sys._MEIPASS
    except AttributeError:
        base_path = os.path.abspath(".")

    return os.path.join(base_path, relative_path)

# Function to limit input entry length in UI
def limit_chars(var, limit):
    if len(var.get()) > limit:
        var.set(var.get()[:limit])

# Function to limit characters in a Text widget (e.g., Address)
def limit_text_chars(event, text_widget, limit=120):
    current_text = text_widget.get("1.0", "end-1c")  # Get text without newline at the end
    if len(current_text) > limit:
        text_widget.delete("1.0", "end")  # Clear text
        text_widget.insert("1.0", current_text[:limit])

# Function to get dropdown values from the Excel sheet
def get_dropdown_values(file_path, sheet_name, column_name):
    workbook = openpyxl.load_workbook(file_path, data_only=True)  # Load Excel
    sheet = workbook[sheet_name]  # Select sheet

     # Find column index by header name
    headers = [cell.value for cell in sheet[1]]  # Read header row (1st row)
    if column_name not in headers:
        print(f"Error: Column '{column_name}' not found in sheet '{sheet_name}'")
        return []
    
    col_index = headers.index(column_name) + 1  # Convert to 1-based index

    # Extract values from the column (starting from row 2)
    values = [
        sheet.cell(row=i, column=col_index).value
        for i in range(2, sheet.max_row + 1)
        if sheet.cell(row=i, column=col_index).value is not None
    ]
    return values

# Tkinter Window
root = tk.Tk()
root.configure(bg="white")
root.title("Receipt App")
root.geometry(f"{root.winfo_screenwidth()}x{root.winfo_screenheight()}")

# Validation functions
validate_number = root.register(lambda char: char.isdigit() or char == "")
validate_alpha = root.register(lambda char: char.isalpha() or char.isspace() or char == "")

# Main Frame (Left)
main_frame = tk.Frame(root)
main_frame.grid(row=0, column=0, padx=20, pady=20)

# Payment Selection
selection_var = tk.StringVar(value="Cash")
selection_entity = tk.StringVar(value='SGPM_DN')
selected_indx = 0 
checkbox_var = tk.IntVar(value=0)
selected_heading = tk.StringVar()
pdf_file = tk.StringVar()
pdf_file.set(pdf_headings[0])

excel_path = resource_path(entity_xcls[0])
file_name = os.path.basename(excel_path)
dest_excel_path = f"excel/{file_name}"
if not os.path.exists(dest_excel_path):
    os.makedirs(os.path.dirname(dest_excel_path), exist_ok=True)
    shutil.copy(excel_path, dest_excel_path)
    print(f"Copied {excel_path} to {dest_excel_path}")
else:
    print(f"Excel File already exists.")
    
contribution_types = get_dropdown_values(dest_excel_path, mapping, "Contribution Type")  # Load from "ContributionType" sheet
contribution_intents = get_dropdown_values(dest_excel_path, mapping, "Contribution Intent")  # Load from "ContributionIntent" sheet

# Define variables to store selected values
contribution_type_var = tk.StringVar(value=contribution_types[0] if contribution_types else "")
contribution_intent_var = tk.StringVar(value=contribution_intents[0] if contribution_intents else "")

# Creating Frames to group fields with borders
frame_positions = [(0, 0), (1, 0), (2, 0), (3, 0), (0, 3)]          
group_frames = {}
for i, (row, col) in enumerate(frame_positions, start=1):
    frame_name = f"group{i}_frame"
    frame = tk.Frame(root, borderwidth=2, relief="groove", bg="#99ccff")
    frame.grid(row=row, column=col, padx=20, pady=20, sticky="nsew")
    group_frames[frame_name] = frame

# Grouped Fields (you can adjust as needed)
fields_group_1 = [
    ("Receipt Date:", DateEntry, {}),
    ("Contributor Name:", tk.Entry, {"textvariable": tk.StringVar()}),
    ("Address:", tk.Text, {"height": 3}),
    ("PAN:", tk.Entry, {"textvariable": tk.StringVar()}),
]

fields_group_2 = [
    ("Contribution Type:", ttk.Combobox, {"values": contribution_types, "textvariable": contribution_type_var}),
    ("Contribution Intent:", ttk.Combobox, {"values": contribution_intents, "textvariable": contribution_intent_var}),
]

fields_group_3 = [
    ("Bank name:", tk.Entry, {"validate": "key", "validatecommand": (validate_alpha, "%S")}),
    ("Branch:", tk.Entry, {"validate": "key", "validatecommand": (validate_alpha, "%S")}),
    ("Amount:", tk.Entry, {"validate": "key", "validatecommand": (validate_number, "%S")}),
]

fields_group_4 = []
fields_group_5 = []

# Entity Selection
selection_entity.set(headings[0]) 
tk.Label(group_frames["group5_frame"], text="Select the Entity:", font=font_settings, bg="#99ccff", relief="groove").grid(row=0, column=3, sticky="w", pady=5)
for idx, option in enumerate(headings):
    radiobutton = tk.Radiobutton(group_frames["group5_frame"], text=option, variable=selection_entity, value=option, font=font_settings, bg="#99ccff") 
    radiobutton.grid(row=idx + 1, column=3, sticky="w")


input_vars = {}                                                     # Store variables for later access
# Function to add widgets dynamically to each group
def add_widgets_to_group(frame, fields):
    for i, (label, widget, field_options) in enumerate(fields):
        tk.Label(frame, text=label, font=font_settings, bg="#99ccff").grid(row=i, column=0, sticky="w", pady=5)
        
        # Get the variable, if any, from field_options
        var = field_options.get("textvariable", None)
        
        # Check if a variable exists for the field (e.g., StringVar for entry fields)
        if var:
            limit = 10 if "PAN:" in label else 27 if "UTRN:" in label else 60
            var.trace_add("write", lambda *args, v=var, l=limit: limit_chars(v, l))
            input_vars[label] = var

        # Handle OptionMenu separately, since it requires both 'variable' and 'value' parameters
        if widget == tk.OptionMenu:
            var = field_options.get("variable")  # Ensure the 'variable' is passed
            options = field_options.get("options")  # Ensure the options are passed
            if var is not None and options is not None:
                entry = widget(frame, var, *options)
            else:
                raise ValueError(f"OptionMenu requires both 'variable' and 'options' in field_options. Missing for '{label}'")
        elif label == 'Receipt Date:':
            entry = widget(frame, font=font_settings, width=5, **field_options)
        else:
            entry = widget(frame, font=font_settings, width=width, **field_options)

        # Add the widget to the grid
        entry.grid(row=i, column=1, pady=5, sticky="ew")
        input_vars[label] = entry

# Add widgets to the frames for each group of fields
field_groups = {
    "group1_frame": fields_group_1, "group2_frame": fields_group_2, "group3_frame": fields_group_3,
    "group4_frame": fields_group_4, "group5_frame": fields_group_5,
}
for group_name, frame in group_frames.items():
    fields = field_groups.get(group_name)
    if fields:
        add_widgets_to_group(frame, fields)

# Payment Mode Selection
tk.Label(group_frames["group4_frame"], text="Select Payment Mode:", font=font_settings, bg="#99ccff").grid(row=3, column=0, sticky="w", pady=5)
cash = tk.Radiobutton(group_frames["group4_frame"], text="Cash", bg="#99ccff", variable=selection_var, value="Cash", font=font_settings, command=lambda: toggle_cheque_fields())
cheque = tk.Radiobutton(group_frames["group4_frame"], text="Cheque", bg="#99ccff", variable=selection_var, value="Cheque", font=font_settings, command=lambda: toggle_cheque_fields())
online = tk.Radiobutton(group_frames["group4_frame"],  text="EFT", bg="#99ccff", variable=selection_var, value="EFT", font=font_settings, command=lambda: toggle_cheque_fields())
cash.grid(row=3, column=1, sticky="w")
cheque.grid(row=3, column=2, sticky="w")
online.grid(row=3, column=3, sticky="w")

checkbox = tk.Checkbutton(group_frames["group4_frame"], text="Include background", bg="#99ccff", variable=checkbox_var)
checkbox.grid(row=6, column=0, sticky="w")

# Cheque Frame (Right)
cheque_frame = tk.Frame(root, borderwidth=2, relief="groove")
cheque_frame.grid(row=0, column=1, padx=50, pady=20, sticky="n")
cheque_frame.grid_remove()  # Hide initially

# Online Frame (Right)
online_frame = tk.Frame(root, borderwidth=2, relief="groove")
online_frame.grid(row=0, column=1, padx=50, pady=20, sticky="n")
online_frame.grid_remove()  # Hide initially

tk.Label(cheque_frame, text="Cheque Details", font=(font_settings[0], font_settings[1], "bold")).grid(row=0, column=0, columnspan=2, pady=10)

cheque_fields = [
    ("Cheque Date:", DateEntry, {}),
    ("Cheque No.:", tk.Entry, {"textvariable": tk.StringVar(), "validate": "key", "validatecommand": (validate_number, "%P")}),
    ("IFSC Code:", tk.Entry, {"textvariable": tk.StringVar()}),
    ("Account No.:", tk.Entry, {"textvariable": tk.StringVar()})
]

online_fields = [
    ("Bank Date:", DateEntry, {}),
    ("UTRN:", tk.Entry, {"textvariable": tk.StringVar()}),
]

cheque_vars = {}  # Store variables for later access
for i, (label, widget, options) in enumerate(cheque_fields):
    tk.Label(cheque_frame, text=label, font=font_settings).grid(row=i + 1, column=0, padx=5, pady=5, sticky="w")
    
    var = options.get("textvariable", None)
    if var:  # Apply character limits
        limit = 6 if "Cheque No" in label else 11 if "IFSC Code" in label else 12
        var.trace_add("write", lambda *args, v=var, l=limit: limit_chars(v, l))
        cheque_vars[label] = var  # Store for later
    
    entry = widget(cheque_frame, font=font_settings, width=13, **options)
    entry.grid(row=i + 1, column=1, padx=5, pady=5)
    cheque_vars[label] = entry  # Store reference

online_vars = {}
for i, (label, widget, options) in enumerate(online_fields):
    tk.Label(online_frame, text=label, font=font_settings).grid(row=i + 1, column=0, padx=5, pady=5, sticky="w")
    
    var = options.get("textvariable", None)
    if var:  # Apply character limits
        if "UTRN:" in label:
            limit = 27 
        var.trace_add("write", lambda *args, v=var, l=limit: limit_chars(v, l))
        online_vars[label] = var  # Store for later
    
    entry = widget(online_frame, font=font_settings, width=28, **options)
    entry.grid(row=i + 1, column=1, padx=5, pady=5)
    online_vars[label] = entry  # Store reference

# Submit Function
def submit():   
    details = {}
    selected_value = selection_entity.get()  # Get the value of the selected radiobutton
    if selected_value in headings:
        selected_indx = headings.index(selected_value)  # Get the index of the selected value
    else:
        print("No option selected.")

    for label, var in input_vars.items():
        if isinstance(var, tk.StringVar):                       
            details[label] = var.get()
        elif isinstance(var, tk.Entry):                         # Normal Entry widget
            details[label] = var.get()
        elif isinstance(var, tk.Text):                          # Multiline Text widget
            details[label] = var.get("1.0", "end-1c").strip()   # Remove extra newlines
        else:
            details[label] = ""                                 # Handle unexpected types safely
    #print("Details:", details)
    
    if selection_var.get() == "Cheque":
        cheque_details = {
            label: var.get() if isinstance(var, tk.StringVar) 
            else var.get("1.0", "end-1c") if isinstance(var, tk.Text)
            else var.get()
            for label, var in cheque_vars.items()}
        payment_mode = 'Cheque'
        cheque_date = cheque_details.get('Cheque Date:')
        cheque_no = cheque_details.get('Cheque No.:')
        IFSC = cheque_details.get('IFSC Code:')
        ac_no =  cheque_details.get('Account No.:')
        bank_date, utrn = '', ''
    elif selection_var.get() in ("Cash", "EFT"):
        payment_mode = selection_var.get()
        cheque_date, cheque_no, IFSC, ac_no = '', '', '', ''
        if payment_mode == "EFT":
            online_details = {
                label: var.get() if isinstance(var, tk.StringVar) 
                else var.get("1.0", "end-1c") if isinstance(var, tk.Text)
                else var.get()
                for label, var in online_vars.items()}
            bank_date = online_details.get('Bank Date:')
            utrn = online_details.get('UTRN:')
        elif payment_mode == 'Cash':
            bank_date, utrn = '', ''
    
    #index = selected_index.get()
    print(f"Index is {selected_indx}")
    entity_path = resource_path(entity_xcls[selected_indx])
    file_name = os.path.basename(entity_path)
    dest_excel_path = f"excel/{file_name}"
    if not os.path.exists(dest_excel_path):
        os.makedirs(os.path.dirname(dest_excel_path), exist_ok=True)
        shutil.copy(entity_path, dest_excel_path)
        print(f"Copied new {entity_path} to {dest_excel_path}")
    else:
        print(f"File {dest_excel_path} already exists.")
    #id = excel_data.increment_counter(1, dest_excel_path)
    id, globe_id, bar_text = excel_data.increment_counter(dest_excel_path)
    print (f"Globe ID {globe_id} TextColumn {bar_text}")

    date = datetime.datetime.now()
    formatted_date = date.strftime("%m/%d/%Y %H.%M.%S")

    barcode_path = excel_data.generate_barcode(str(globe_id))   #barcode is temporarily saved in the current dir
    excel_data.draw_text(f"{barcode_path}.png", bar_text)       #adding TextColumn to the barcode

    #the col header and order here should be an exact match with the excel.
    kwargs = {
        'Id.' : id,
        'Receipt No.' : globe_id, 
        'Receipt Date' : details.get("Receipt Date:"),
        'Amount' : details.get("Amount:"),
        'Contributor Name' : details.get("Contributor Name:"),
        'Contributor Type' : details.get("Contribution Type:"),
        'Contributor Intent' : details.get("Contribution Intent:"),
        'Bank Date' : bank_date,
        'UTR No.' : utrn,
        'Cheque Date' : cheque_date,
        'Cheque No.' : cheque_no,
        'IFSC' : IFSC,
        'Account No.' : ac_no,
        'Address' : details.get("Address:"),
        'PAN' : details.get("PAN:"),
        'Payment Mode' : payment_mode,
        'Barcode' : f"{barcode_path}.png",
        'Bank Name' : details.get("Bank name:"),
        'Branch Name' : details.get("Branch:"),
        'Print Date' : formatted_date,
        'Globe Id.' : globe_id,
        'TextColumn' : bar_text,
        'QR Code' : '',
    }

    excel_data.save_to_excel(dest_excel_path, id, **kwargs)

    qr_code_path = excel_data.generate_qr_code(kwargs)
    kwargs["QR Code"] = qr_code_path       
    entity_name = os.path.splitext(entity[selected_indx])[0]                 # Update kwargs with QR Code path
    pdf_path = f"pdfs/{entity_name}/{kwargs['Id.']}.pdf"
    os.makedirs(f"pdfs/{entity_name}", exist_ok=True)
    pdf_data.create_pdf_from_kwargs(kwargs, pdf_path, entity[selected_indx], checkbox_var.get(), copy_types[1], copy_types[2])
    pdf_data.create_pdf_from_kwargs(kwargs, pdf_path, entity[selected_indx], checkbox_var.get(), copy_types[0], copy_types[0])

    try:
        df = pd.read_excel(dest_excel_path, sheet_name=xcl_sheet)
        workbook = openpyxl.load_workbook(dest_excel_path)
        sheet = workbook[xcl_sheet] 
        pdf_dir = pdf_data.get_pdf_directory()
        os.makedirs(pdf_dir, exist_ok=True)
    except ValueError as e:
        print(f"Error: Worksheet named '{xcl_sheet}' not found. Please check the sheet name.")
        messagebox.showinfo(f"Error:", f"Worksheet named '{xcl_sheet}' not found. Please check the sheet name.") 
    except Exception as e:
        print(e)

    rows_data = []                                          # Loop through each row in the DataFrame and create a separate PDF

    for index, row in df.iterrows():
        row_dict = {df.columns[i]: row.iloc[i] for i in range(len(df.columns))}
        rows_data.insert(id, row_dict) 
        pdf_name = f"pdfs/{index}.pdf"
    workbook.save(dest_excel_path)

submit_button = tk.Button(group_frames["group4_frame"], text="Print", font=font_settings, command=submit)   # Submit Button
submit_button.grid(row=6, column=1, columnspan=2, pady=20)

# Toggle Cheque Fields
def toggle_cheque_fields():
    cheque_frame.grid() if selection_var.get() == "Cheque" else cheque_frame.grid_remove()
    online_frame.grid() if selection_var.get() == "EFT" else online_frame.grid_remove()

toggle_cheque_fields()                                                              # Ensure correct visibility
root.mainloop()
