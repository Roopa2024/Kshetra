import tkinter as tk
from tkinter import messagebox
import pandas as pd
from openpyxl import load_workbook
import re, os
import backend.shared
from openpyxl import Workbook

file_name = backend.shared.xcl_file 
sheet_name = backend.shared.xcl_sheet 

def increment_counter():
    if not os.path.exists(file_name):
        print(f"⚠️ File '{file_name}' not found. Creating a new one...")
        wb = Workbook()  # Create a new workbook
        wb.active.append(["Counter"])  # Add a header row (optional)
        wb.save(file_name)  # Save new workbook
        print("✅ New file created successfully!")

    wb = load_workbook(file_name)   # Load the workbook and select the sheet
    sheet = wb.active               # Or use wb['SheetName'] if you have a specific sheet
    last_row = sheet.max_row        # Get the last row number
    last_value = sheet.cell(row=last_row, column=2).value  # Adjust column index as needed
    print(f"Value is {last_value}")
    match = re.match(r'([a-zA-Z]+)(\d+)', last_value)
    
    if match:
        prefix = match.group(1)     # Extract the prefix and the number part
        num_part = match.group(2)
        
        # Increment the numeric part and pad it to 4 digits (e.g., 0002)
        num_part = str(int(num_part) + 1).zfill(len(num_part))  # Use zfill to maintain leading zeros
        
        new_value = prefix + num_part # Reassemble the string with the incremented value
        print(f"New value: {new_value}")
        return new_value
    else:
        print("Invalid format")
        return None

def save_to_excel(**kwargs):
    for key, value in kwargs.items():
        print(f"{key}: {value}")
    data = {key: [value] for key, value in kwargs.items()} # Convert each value in kwargs to a list
      
    df = pd.DataFrame(data) # Convert dictionary to pandas DataFrame
    
    # If the file exists, append the new data to it, else create a new file
    try:
        existing_df = pd.read_excel(file_name)         # Check if file exists
        new_df = pd.concat([existing_df, df], ignore_index=True)
        new_df.to_excel(file_name, index=False)
    except FileNotFoundError:
        df.to_excel(file_name, index=False)         # Create a new Excel file if it does not exist

    messagebox.showinfo("Success", "Data saved successfully to Excel!")

def load_column_values_to_dropdown(column_name):
    df = pd.read_excel(file_name, sheet_name)      # Read the spreadsheet and load the column values
    if column_name not in df.columns:     # Ensure the column exists in the DataFrame
        raise ValueError(f"Column '{column_name}' not found in the sheet '{sheet_name}'")
        
    dropdown_values = df[column_name].dropna().tolist()  # Drop any NaN values and convert to list
    return dropdown_values
